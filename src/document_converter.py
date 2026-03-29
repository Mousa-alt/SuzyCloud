"""
Document format converter — LibreOffice headless PDF conversion with ODS intermediate fallback.

Converts existing documents (.xlsx, .docx, .pptx, etc.) to PDF.
NOT for generating documents from text — use document_gen.py for that.

LibreOffice is the primary converter. For spreadsheet files (.xlsx, .xls, .csv, .ods),
if LibreOffice hangs or crashes (known issue with large/complex workbooks), the converter
automatically falls back to a multi-step pipeline:
  1. ssconvert (gnumeric) converts the spreadsheet to ODS format
  2. ODS XML is post-processed to restore page setup (scale percentages, master pages, paper size)
  3. LibreOffice converts the fixed ODS to PDF

Images (including EMF/WMF) are left untouched — LibreOffice handles them natively during
PDF export, preserving their original proportions and scaling them with the page layout.

Usage as CLI:
    python src/document_converter.py <input_file> [output_dir] [--a3|--a4] [--portrait|--landscape] [--fit-columns]

    Options (Excel files only):
        --a3            A3 paper size (default: unchanged)
        --a4            A4 paper size
        --portrait      Portrait orientation
        --landscape     Landscape orientation
        --fit-columns   Fit all columns to one page width

Usage as module:
    from src.document_converter import convert_to_pdf
    pdf_path = convert_to_pdf("media/incoming/document/file.xlsx")
    pdf_path = convert_to_pdf("file.xlsx", page_setup={"paper": "a3", "orientation": "portrait", "fit_columns": True})

Supported: .xlsx, .xls, .docx, .doc, .pptx, .ppt, .csv, .odt, .ods, .odp, .txt, .rtf
"""

import argparse
import logging
import os
import signal
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

log = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {
    ".xlsx", ".xls", ".docx", ".doc", ".pptx", ".ppt",
    ".csv", ".odt", ".ods", ".odp", ".txt", ".rtf",
}

EXCEL_EXTENSIONS = {".xlsx"}

# Spreadsheet formats that the ODS intermediate pipeline can handle
SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv", ".ods"}

# openpyxl paper size constants
PAPER_SIZES = {
    "a4": 9,
    "a3": 8,
    "letter": 1,
}

TIMEOUT_NORMAL = 120
TIMEOUT_LARGE = 300
# Shorter LO timeout for spreadsheets when ODS fallback is available
TIMEOUT_SPREADSHEET = 60
SSCONVERT_TIMEOUT = 300
LARGE_FILE_THRESHOLD = 5 * 1024 * 1024  # 5 MB

# ODS XML namespaces (used by _fix_ods_page_setup)
_ODS_NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "style": "urn:oasis:names:tc:opendocument:xmlns:style:1.0",
    "fo": "urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0",
    "gnm": "http://www.gnumeric.org/odf-extension/1.0",
}


def _apply_excel_page_setup(src_path: Path, page_setup: dict, tmp_dir: str) -> Path:
    """Apply page setup to all sheets of an Excel file, return path to modified copy."""
    import openpyxl
    from openpyxl.worksheet.properties import PageSetupProperties

    wb = openpyxl.load_workbook(str(src_path))

    paper = page_setup.get("paper")
    orientation = page_setup.get("orientation")
    fit_columns = page_setup.get("fit_columns", False)

    normalized_paper = paper.lower() if isinstance(paper, str) else None
    normalized_orientation = orientation.lower() if isinstance(orientation, str) else None

    if normalized_paper and normalized_paper not in PAPER_SIZES:
        wb.close()
        raise ValueError(f"Unsupported paper size: {paper}")
    if normalized_orientation and normalized_orientation not in {"portrait", "landscape"}:
        wb.close()
        raise ValueError(f"Unsupported orientation: {orientation}")

    for ws in wb.worksheets:
        if normalized_paper:
            ws.page_setup.paperSize = PAPER_SIZES[normalized_paper]
        if normalized_orientation:
            ws.page_setup.orientation = normalized_orientation
        if fit_columns:
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    modified = Path(tmp_dir) / src_path.name
    wb.save(str(modified))
    wb.close()
    return modified


def _convert_with_libreoffice(convert_src: Path, tmp_dir: str, timeout: int) -> Path | None:
    """Try converting with LibreOffice. Returns PDF path on success, None on failure."""
    # Remove any stale PDF from a prior timed-out run in the same tmp_dir
    expected_pdf = Path(tmp_dir) / (convert_src.stem + ".pdf")
    if expected_pdf.exists():
        try:
            expected_pdf.unlink()
        except OSError as e:
            log.warning("Could not remove stale PDF %s: %s", expected_pdf, e)

    cmd = [
        "libreoffice",
        "--headless",
        "--norestore",
        "--convert-to", "pdf",
        "--outdir", tmp_dir,
        str(convert_src),
    ]

    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            start_new_session=True,
        )
        stdout, stderr = proc.communicate(timeout=timeout)
    except FileNotFoundError:
        log.warning("LibreOffice not found on PATH")
        return None
    except subprocess.TimeoutExpired:
        log.warning("LibreOffice timed out after %ds", timeout)
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except (ProcessLookupError, OSError):
            proc.kill()
        proc.wait()
        return None

    if expected_pdf.exists() and expected_pdf.stat().st_size > 0:
        return expected_pdf

    # Check for "source file could not be loaded" or other LO errors
    combined = (stdout or "") + (stderr or "")
    if combined.strip():
        log.warning("LibreOffice produced no PDF. Output: %s", combined.strip()[:300])
    else:
        log.warning("LibreOffice exited code %d with no output", proc.returncode)
    return None


def _read_xlsx_scales(xlsx_path: Path) -> list[tuple[str, int]]:
    """Read per-sheet scale percentages from an Excel file using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    scales = []
    for name in wb.sheetnames:
        ws = wb[name]
        scale = int(ws.page_setup.scale) if ws.page_setup.scale else 100
        scales.append((name, scale))
    wb.close()
    return scales



_ODS_PAPER_DIMS = {
    "a3": ("29.7cm", "42cm"),
    "a4": ("21cm", "29.7cm"),
    "letter": ("21.59cm", "27.94cm"),
}


def _fix_ods_page_setup(ods_path: Path, sheet_scales: list[tuple[str, int]],
                        tmp_dir: str,
                        page_setup: dict | None = None) -> Path:
    """Fix ssconvert ODS output to restore page setup from the original xlsx.

    ssconvert drops two critical pieces of page setup when converting to ODS:
    1. master-page-name on table styles (breaks the style chain: table -> master page -> page layout)
    2. scale-to percentages (uses gnumeric-specific attributes instead of standard ODF ones)

    If page_setup is provided, paper size, orientation, and fit-columns are applied
    to all page layouts in the ODS XML.

    Images are NOT touched — LibreOffice handles EMF/WMF natively during PDF export.

    This reads the ODS ZIP, patches the XML, and writes a fixed copy.
    """
    from lxml import etree

    style_ns = _ODS_NS["style"]
    gnm_ns = _ODS_NS["gnm"]

    fixed_path = Path(tmp_dir) / f"{ods_path.stem}_fixed.ods"

    with zipfile.ZipFile(str(ods_path), "r") as zin:
        content_xml = zin.read("content.xml")
        styles_xml = zin.read("styles.xml")

        content_tree = etree.fromstring(content_xml)
        styles_tree = etree.fromstring(styles_xml)

        # Fix 1: Set master-page-name on each table style.
        # ssconvert sets master-page=None, breaking the table → master page link.
        auto_styles = content_tree.find(f"{{{_ODS_NS['office']}}}automatic-styles")
        fixed_tables = 0
        if auto_styles is not None:
            for style_el in auto_styles:
                sname = style_el.get(f"{{{style_ns}}}name")
                sfamily = style_el.get(f"{{{style_ns}}}family")
                if sfamily == "table" and sname and sname.startswith("ta-"):
                    try:
                        idx = int(sname.split("-")[1])
                    except (IndexError, ValueError):
                        continue
                    style_el.set(f"{{{style_ns}}}master-page-name", f"ta-mp-{idx}")
                    fixed_tables += 1

        # Fix 2: Set standard scale-to on each page layout.
        # ssconvert uses {gnumeric}scale-to-X instead of the standard {style}scale-to.
        page_layouts = styles_tree.findall(f".//{{{style_ns}}}page-layout")
        fixed_scales = 0
        for pl in page_layouts:
            pl_name = pl.get(f"{{{style_ns}}}name")
            if not (pl_name and pl_name.startswith("pl-")):
                continue
            try:
                idx = int(pl_name.split("-")[1])
            except (IndexError, ValueError):
                continue
            if idx >= len(sheet_scales):
                continue
            _name, scale = sheet_scales[idx]
            props = pl.find(f"{{{style_ns}}}page-layout-properties")
            if props is None:
                continue
            props.set(f"{{{style_ns}}}scale-to", str(scale))
            # Remove gnumeric-specific attribute that conflicts
            gnm_attr = f"{{{gnm_ns}}}scale-to-X"
            if gnm_attr in props.attrib:
                del props.attrib[gnm_attr]
            fixed_scales += 1

        log.info("ODS fix: %d table styles, %d scale values patched", fixed_tables, fixed_scales)

        # Fix 2b: Apply page_setup overrides (paper size, orientation, fit-columns).
        # This replaces the openpyxl step which can't be used because it drops EMF images.
        if page_setup:
            fo_ns = _ODS_NS["fo"]
            paper = (page_setup.get("paper") or "").lower()
            orientation = (page_setup.get("orientation") or "").lower()
            fit_columns = page_setup.get("fit_columns", False)
            dims = _ODS_PAPER_DIMS.get(paper)

            for pl in page_layouts:
                props = pl.find(f"{{{style_ns}}}page-layout-properties")
                if props is None:
                    continue
                if dims:
                    # Use explicit orientation, or preserve existing orientation
                    effective_orient = orientation or props.get(
                        f"{{{style_ns}}}print-orientation", "portrait"
                    )
                    if effective_orient == "landscape":
                        props.set(f"{{{fo_ns}}}page-width", dims[1])
                        props.set(f"{{{fo_ns}}}page-height", dims[0])
                    else:
                        props.set(f"{{{fo_ns}}}page-width", dims[0])
                        props.set(f"{{{fo_ns}}}page-height", dims[1])
                if orientation:
                    props.set(f"{{{style_ns}}}print-orientation", orientation)
                if fit_columns:
                    # Remove gnumeric-specific scale-to-X/Y attributes that
                    # LibreOffice ignores. Per-sheet scale percentages from
                    # the original xlsx are already applied via Fix 2 above.
                    for gnm_attr in (f"{{{gnm_ns}}}scale-to-X", f"{{{gnm_ns}}}scale-to-Y"):
                        if gnm_attr in props.attrib:
                            del props.attrib[gnm_attr]
            log.info("ODS fix: page_setup applied (paper=%s, orient=%s, fit=%s)",
                     paper or "unchanged", orientation or "unchanged", fit_columns)

        # Write fixed ODS (only styles.xml and content.xml are modified)
        with zipfile.ZipFile(str(fixed_path), "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "styles.xml":
                    zout.writestr(item, etree.tostring(styles_tree, xml_declaration=True, encoding="UTF-8"))
                elif item.filename == "content.xml":
                    zout.writestr(item, etree.tostring(content_tree, xml_declaration=True, encoding="UTF-8"))
                else:
                    zout.writestr(item, zin.read(item.filename))

    return fixed_path


def _convert_via_ods_intermediate(src: Path, tmp_dir: str,
                                   page_setup: dict | None = None) -> Path | None:
    """Convert a spreadsheet to PDF via ODS intermediate format.

    Pipeline: xlsx/xls/csv -> ssconvert -> ODS -> fix page setup XML -> LibreOffice -> PDF

    This avoids LibreOffice crashing on complex XLSX files while preserving
    page layout (scale percentages, paper size, orientation) and embedded images.
    Images (including EMF/WMF) are left as-is — LibreOffice renders them natively.
    """
    # Step 1: Read scale percentages from original file (xlsx only)
    sheet_scales: list[tuple[str, int]] = []
    if src.suffix.lower() == ".xlsx":
        try:
            sheet_scales = _read_xlsx_scales(src)
            log.info("Read %d sheet scales from %s", len(sheet_scales), src.name)
        except Exception as e:
            log.warning("Could not read scales from %s: %s (%s)", src.name, e, type(e).__name__)

    # Step 2: Convert to ODS via ssconvert
    ods_path = Path(tmp_dir) / (src.stem + ".ods")
    cmd = ["ssconvert", str(src), str(ods_path)]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=SSCONVERT_TIMEOUT,
        )
    except FileNotFoundError:
        log.warning("ssconvert (gnumeric) not found on PATH")
        return None
    except subprocess.TimeoutExpired:
        log.warning("ssconvert xlsx->ods timed out after %ds", SSCONVERT_TIMEOUT)
        return None

    if result.returncode != 0:
        log.warning("ssconvert xlsx->ods failed (exit %d): %s",
                     result.returncode, result.stderr.strip()[:300])
        return None

    if not (ods_path.exists() and ods_path.stat().st_size > 0):
        log.warning("ssconvert xlsx->ods produced no output file")
        return None

    log.info("ssconvert produced ODS: %d bytes", ods_path.stat().st_size)

    # Step 3: Fix ODS page setup (scale percentages, paper size, orientation)
    convert_ods = ods_path
    if sheet_scales or page_setup:
        try:
            convert_ods = _fix_ods_page_setup(
                ods_path, sheet_scales, tmp_dir, page_setup)
        except Exception as e:
            log.warning("ODS fix failed, using unfixed ODS: %s (%s)", e, type(e).__name__)
            convert_ods = ods_path

    # Step 4: Convert ODS to PDF via LibreOffice
    pdf_path = _convert_with_libreoffice(convert_ods, tmp_dir, TIMEOUT_LARGE)
    if pdf_path is None:
        log.warning("LibreOffice failed to convert ODS intermediate file")
    return pdf_path


def convert_to_pdf(
    input_path: str,
    output_dir: str = "media/outgoing",
    page_setup: dict | None = None,
    output_name: str | None = None,
) -> str:
    """Convert a document to PDF using LibreOffice headless, with ODS intermediate fallback.

    For spreadsheet files (.xlsx, .xls, .csv, .ods), if LibreOffice fails or times out,
    automatically falls back to a pipeline: ssconvert → ODS → fix page setup → LO → PDF.

    Args:
        input_path: Path to the source file.
        output_dir: Directory for the output PDF.
        page_setup: Optional dict with Excel page settings:
            - paper: "a3", "a4", or "letter"
            - orientation: "portrait" or "landscape"
            - fit_columns: True to fit all columns to one page width
        output_name: Optional base name for the output PDF (without .pdf extension).
            If provided, the PDF will be named "{output_name}.pdf" instead of
            using the input file's stem. Useful when the input file has an
            auto-generated name (e.g. Waha media ID) but the original
            document name is known.

    Returns:
        Relative path to the generated PDF (e.g. media/outgoing/file.pdf).

    Raises:
        FileNotFoundError: Input file doesn't exist.
        ValueError: Unsupported file format.
        RuntimeError: Conversion failed or timed out.
    """
    src = Path(input_path).resolve()
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if not src.is_file():
        raise ValueError(f"Input path is not a file: {input_path}")

    ext = src.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported format '{ext}'. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    out_dir = Path(output_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    is_large = src.stat().st_size > LARGE_FILE_THRESHOLD
    timeout = TIMEOUT_LARGE if is_large else TIMEOUT_NORMAL

    with tempfile.TemporaryDirectory(prefix="lo_convert_") as tmp_dir:
        convert_src = src

        # Check if ODS intermediate fallback is available
        has_ods_fallback = ext in SPREADSHEET_EXTENSIONS and shutil.which("ssconvert") is not None

        # When page_setup is requested and ODS fallback is available, go directly
        # to the ODS pipeline. This avoids openpyxl (which drops EMF/WMF images
        # on save) and ensures page setup is applied safely in ODS XML.
        if page_setup and has_ods_fallback and ext in EXCEL_EXTENSIONS:
            log.info("Using ODS pipeline for %s (page_setup requested, preserves images)", src.name)
            pdf_path = _convert_via_ods_intermediate(src, tmp_dir, page_setup=page_setup)
        elif page_setup and ext == ".xls":
            raise ValueError(
                "Page setup flags require .xlsx format. "
                "Please save the workbook as .xlsx and retry."
            )
        else:
            # Apply page setup via openpyxl only when no ODS fallback available
            if page_setup and ext in EXCEL_EXTENSIONS:
                convert_src = _apply_excel_page_setup(src, page_setup, tmp_dir)
                timeout = max(timeout, TIMEOUT_LARGE)

            lo_timeout = min(timeout, TIMEOUT_SPREADSHEET) if has_ods_fallback else timeout

            # Try LibreOffice first (direct conversion)
            pdf_path = _convert_with_libreoffice(convert_src, tmp_dir, lo_timeout)

            # Fallback: ODS intermediate pipeline for spreadsheet formats
            if pdf_path is None and has_ods_fallback:
                log.info("Falling back to ODS intermediate pipeline for %s", src.name)
                pdf_path = _convert_via_ods_intermediate(src, tmp_dir)

        if pdf_path is not None:
            # Use output_name if provided, otherwise fall back to input file stem
            base_name = output_name if output_name else src.stem
            # Sanitize: remove characters unsafe for filenames
            base_name = "".join(c for c in base_name if c not in r'<>:"/\|?*').strip()
            if not base_name:
                base_name = src.stem
            final_name = f"{base_name}.pdf"
            final_pdf = out_dir / final_name
            # Atomically claim the target filename to avoid race conditions
            # when multiple concurrent conversions target the same output dir.
            counter = 0
            while True:
                candidate = final_pdf if counter == 0 else out_dir / f"{base_name}_{counter}.pdf"
                try:
                    fd = os.open(str(candidate), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                    os.close(fd)
                    final_pdf = candidate
                    break
                except FileExistsError:
                    counter += 1
            shutil.move(str(pdf_path), str(final_pdf))
            return f"{output_dir}/{final_pdf.name}"

        if has_ods_fallback:
            raise RuntimeError(
                "Conversion failed. Both LibreOffice and the ODS intermediate "
                "pipeline (ssconvert + LibreOffice) were unable to convert the file."
            )
        raise RuntimeError(
            "Conversion failed. LibreOffice was unable to convert the file."
        )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert documents to PDF via LibreOffice headless."
    )
    parser.add_argument("input_file", help="Path to the source file")
    parser.add_argument("output_dir", nargs="?", default="media/outgoing",
                        help="Output directory (default: media/outgoing)")
    paper_group = parser.add_mutually_exclusive_group()
    paper_group.add_argument("--a3", action="store_const", const="a3", dest="paper",
                             help="A3 paper size (Excel only)")
    paper_group.add_argument("--a4", action="store_const", const="a4", dest="paper",
                             help="A4 paper size (Excel only)")

    orientation_group = parser.add_mutually_exclusive_group()
    orientation_group.add_argument("--portrait", action="store_const", const="portrait",
                                   dest="orientation", help="Portrait orientation (Excel only)")
    orientation_group.add_argument("--landscape", action="store_const", const="landscape",
                                   dest="orientation", help="Landscape orientation (Excel only)")
    parser.add_argument("--fit-columns", action="store_true",
                        help="Fit all columns to one page width (Excel only)")
    parser.add_argument("--output-name", type=str, default=None,
                        help="Base name for the output PDF (without .pdf extension)")

    args = parser.parse_args()

    setup = {}
    if args.paper:
        setup["paper"] = args.paper
    if args.orientation:
        setup["orientation"] = args.orientation
    if args.fit_columns:
        setup["fit_columns"] = True

    try:
        pdf = convert_to_pdf(args.input_file, args.output_dir,
                             page_setup=setup or None,
                             output_name=args.output_name)
        print(pdf)
    except (FileNotFoundError, ValueError, RuntimeError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
